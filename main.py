import traceback
import openpyxl
from package.testip import *
from pathlib import Path

def main():
    current_dir = Path.cwd()
    files = [file for file in current_dir.joinpath('./数据文件/').iterdir() if file.is_file() and '~' not in file.name]
    xlsx_file = files[0]
    try:
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
        
        from package.server import call_run
        changed = False
        while(not changed):
            try:
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=10).value:
                        continue
                    inputs = {}
                    for col in range(1,16):
                        inputs[ws.cell(row=1, column=col).value] = ws.cell(row=row, column=col).value
                    print(inputs['IP地址']+",",end='')
                    result = call_run(inputs)
                    print(result)
                    ws.cell(row=row, column=10).value = result
                    wb.save(xlsx_file)
                    changed = True
            except KeyboardInterrupt:
                input("按键暂停,关闭或按键继续.")
        wb.save(xlsx_file)
        wb.close()
    except:
        traceback.print_exc()
        input("主程序操作失败,按键退出")
    print("全部完成")
    
if __name__ == '__main__':
    main()
    input('按键退出')