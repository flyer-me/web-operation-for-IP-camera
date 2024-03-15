# -*- coding: utf-8 -*-
import sys
from datetime import datetime
from ipaddress import ip_address
from pathlib import Path
from chardet import detect
from openpyxl import load_workbook
import configparser
import traceback
import subprocess
import locale


def listdir(path):
    directory = Path(path)
    return [file.name for file in directory.iterdir() if file.is_file()]

'''
description: 
param {*} filename
param {*} header_row
return {*} max_row
'''
def is_sheet_empty(sheet) ->int:
    return sheet.max_row == 1 and sheet.max_column == 1


'''
description: 尝试从sheet的前n行中找到ip地址所在的列数
param {*} filename
param {*} sheet
return {*} 列数，如果找不到则为-1
'''
def get_ip_row_in_sheet(sheet,n:int=10) ->int:
    if is_sheet_empty(sheet):
        return -1
    max_rows = min(n,sheet.max_row)
    for row in sheet.iter_rows(min_row=1,max_row=max_rows,min_col=1,max_col=sheet.max_column):
        for cell in row:
            try:
                ip_address(cell.value)
                return cell.column
            except ValueError:
                pass
    return -1


'''
description: 
param {*} files
return {*}
'''
def read_xlsx(files=None):
    data = []
    for fn in files:
        try:
            wb = load_workbook(fn, read_only=True)
        except PermissionError:
            input(f"{fn}文件无法读取.")
            return None

        for sheet in wb.worksheets:
            col_num = get_ip_row_in_sheet(sheet)
            if col_num <= -1:
                continue
            for row in sheet.iter_rows():
                cell_value = row[col_num].value
                if cell_value is not None:
                    try:
                        ip_address(cell_value)
                        #if ip_address(cell_value).is_private == False:
                        data.append(cell_value)
                    except ValueError:
                        pass
    data = list(set(data))  # 去重
    return data


def get_ip_list(xfile, result="ip.txt"):
    data = read_xlsx(xfile)
    with open(result, 'w') as f:
        for item in data:
            f.write(f"{item}\n")

'''
description: 调用ping应用程序测试ip,并将结果写回result
param {*} tool_path
param {*} xfile
param {*} result
return {*}
'''
def call_ping(tool_path, xfile, result, timeout = 2000, size = 64):
    if not Path(result).exists():
        with open(result, 'w') as f:
            f.write('')
    path = Path(tool_path).resolve()
    command = f'{path} /loadfile {xfile} /stab {result} /PingTimeout {timeout} /PingSize {size}'

    with open(xfile, 'r') as file:
        count = sum(1 for line in file if line.strip())
    print(f'超时IP数{count}.')
    process = None
    try:
        process = subprocess.Popen(command, shell=True)
        process.communicate()
    except:
        print("ping调用失败.")
        traceback.print_exc()
    finally:
        if process is not None and process.poll() is None:
            process.terminate()
            process.wait()

def get_encoding(file):
    if Path(file).exists():
        with open(file, 'rb') as f:
            data = f.read()
            return detect(data)['encoding']

'''
description: 
param {*} ip_list
param {*} bad_ip_list
return {*}
'''
def get_bad_ip(ip_list, bad_ip_list):
    with open(ip_list, 'r', encoding=get_encoding(ip_list)) as f:
        lines = f.readlines()
    with open(bad_ip_list, 'w') as f:
        for line in lines:
            parts = line.split()
            if parts[1] == '0':
                f.write(parts[0] + '\n')


'''
description: 
param {*} file_list
param {*} bad_ip
return {*}
'''
def write_result(file_list, ip_file, is_in, not_in):
    for fn in file_list:
        try:
            wb = load_workbook(fn)
        except PermissionError:
            print(f"{fn}文件被占用,无法读取.")
            sys.exit(1)

        for sheet in wb.worksheets:
            col_num = get_ip_row_in_sheet(sheet)
            if col_num <= -1:
                continue
            # max_col = sheet.max_column
            # new_col = max_col + 1  # 增加 “是否在线” 统计结果列
            new_col = 3
            sheet.cell(row=1, column=new_col).value = datetime.now().strftime("ping情况%m%d %H:%M")

            with open(ip_file, "r") as f:
                ip_list = set(f.read().splitlines())

            for row in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col_num + 1).value  # sheet.cell 的 column 是从1开始
                try:
                    ip_address(cell_value)
                except ValueError:
                    continue
                if cell_value in ip_list:
                    sheet.cell(row=row, column=new_col).value = is_in
                else:
                    sheet.cell(row=row, column=new_col).value = not_in
        wb.save(fn)

'''
description: 删除本目录下file_ext后缀名的文件
param {*} file_ext 后缀名
'''
def remove_txt(file_ext:str = '.txt'):
    file_path = Path.cwd()
    file_ext = '.txt'
    for path in listdir(file_path):
        path_list = Path(file_path).joinpath(Path(path))
        if path_list.is_file():
            if path_list.suffix == file_ext:
                path_list.unlink()


'''
description: 对XLSX文件列表,执行IP连通性测试,并回写结果
param {*} file_list
param {*} enter 为True时，需要手动输入按键以结束.
return {*}
'''
def ip_xlsx_test(file_list, ping_timeout, ping_size, times):
    ip_list_name = "ip.txt"
    result_name = "result.txt"
    bad_name = "bad.txt"
    ping_relative_path = r'tools\PingInfoView.exe'
    get_ip_list(file_list, ip_list_name)
    # 未测试时bad ip为全部ip
    with open(ip_list_name, 'r') as file1, open(bad_name, 'w') as file2:
        file2.write(file1.read())
    # ping重复操作
    print(f'ping:timeout{ping_timeout}ms,{ping_size}bytes.')
    while(times > 0):
        call_ping(ping_relative_path, bad_name, result_name, ping_timeout, ping_size)
        get_bad_ip(result_name, bad_name)

        times = times - 1
        print(f'less than{times}times ',end='')
    # 结果回写
    write_result(file_list, bad_name, "超时", "已通")
    remove_txt()

def test_ip(relative_path=''):
    print(Path.cwd())
    locale.setlocale(locale.LC_ALL, 'en_US.utf8')
    config = configparser.ConfigParser()
    config.read('config/ping.ini',encoding="utf-8-sig")

    ping_timeout = config.getint('cfg', 'ping_timeout', fallback=2000)    #ping超时时间
    ping_times = config.getint('cfg', 'ping_times', fallback=10)        #ping尝试次数
    ping_size = 64

    try:
        file_list = [relative_path+f for f in listdir(relative_path) if f.endswith('.xlsx')]
        ip_xlsx_test(file_list, ping_timeout, ping_size, ping_times)
    except Exception:
        traceback.print_exc()
        input("发生异常，按任意键退出.")
        remove_txt()

if __name__ == '__main__':
    test_ip(relative_path=".\数据文件\\")
    sys.exit(0)