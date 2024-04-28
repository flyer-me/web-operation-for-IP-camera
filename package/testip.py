# -*- coding: utf-8 -*-
import sys
from datetime import datetime
import IPy
from pathlib import Path
from chardet import detect
from openpyxl import load_workbook
import configparser
import traceback
import subprocess
import locale

"""
This is DEMO code. Do not use it in released project.
flyer-me 2024/04/15.
"""


def listdir(path=''):
    directory = Path(path)
    return [file.name for file in directory.iterdir() if file.is_file()]


def is_sheet_empty(sheet) -> int:
    """
    Check if the given sheet is empty.

    Args:
        sheet: The sheet to check.

    Returns:
        int: 1 if the sheet is empty, 0 otherwise.
    """
    return sheet.max_row == 1 and sheet.max_column == 1


def get_ip_row_in_sheet(sheet, n: int = 10) -> int:
    """
    Returns the column number of the first cell containing a valid IP address in the given sheet.

    Args:
        sheet: The sheet object representing the Excel sheet.
        n (optional): The maximum number of rows to check for IP addresses. Defaults to 10.

    Returns:
        The column number of the first cell containing a valid IP address start from 1, or -1 if no IP address is found.

    """
    if is_sheet_empty(sheet):
        return -1
    max_rows = min(n, sheet.max_row)
    for row in sheet.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=sheet.max_column):
        for cell in row:
            try:
                if '.' in cell.value:
                    ip_version = IPy.IP(cell.value).version()
                    if ip_version == 4 or ip_version == 6:
                        return cell.column
            except Exception as e:
                pass
    return -1


def read_xlsx(files=None):
    """
    Read IP addresses from Excel files.

    Args:
        files (list): List of file paths to Excel files.

    Returns:
        list: List of unique IP addresses extracted from the Excel files.
    """
    data = []
    for fn in files:
        try:
            wb = load_workbook(fn, read_only=True)
        except PermissionError:
            input(f"{fn}文件无法读取.")
            return None

        for sheet in wb.worksheets:
            col_num = get_ip_row_in_sheet(sheet)
            #print(f'IP地址位于{col_num}列')
            if col_num <= -1:
                continue
            for row in sheet.iter_rows():
                cell_value = row[col_num - 1].value
                if cell_value is not None:
                    try:
                        data.append(cell_value)
                    except ValueError:
                        pass
    data = list(set(data))  # 去重
    return data


def get_ip_list(xfile, result="ip.txt"):
    data = read_xlsx(xfile)
    with open(result, 'w') as f:
        for item in data:
            f.write(f"{item}\n")

def call_ping(tool_path, xfile, result, timeout=2000, size=64):
    """
    Executes the ping command using the specified tool path and parameters.

    Args:
        tool_path (str): The path to the ping tool.
        xfile (str): The path to the input file containing IP addresses.
        result (str): The path to the output file to store the ping results.
        timeout (int, optional): The timeout value for each ping request in milliseconds. Defaults to 2000.
        size (int, optional): The size of the ping request packet in bytes. Defaults to 64.
    """
    if not Path(result).exists():
        with open(result, 'w') as f:
            f.write('')
    path = Path(tool_path).resolve()
    command = f'{path} /loadfile {xfile} /stab {result} /PingTimeout {timeout} /PingSize {size}'

    with open(xfile, 'r') as file:
        count = sum(1 for line in file if line.strip())
    print(f'{count}IP超时.',end='\r')
    sys.stdout.flush()
    process = None
    try:
        process = subprocess.Popen(command, shell=True)
        process.communicate()
    except:
        print("ping调用失败.")
        traceback.print_exc()
        return False
    finally:
        if process is not None and process.poll() is None:
            process.terminate()
            process.wait()
    return True

def get_encoding(file):
    if Path(file).exists():
        with open(file, 'rb') as f:
            data = f.read()
            return detect(data)['encoding']


def get_bad_ip(ip_list, bad_ip_list):
    with open(ip_list, 'r', encoding=get_encoding(ip_list)) as f:
        lines = f.readlines()
    with open(bad_ip_list, 'w') as f:
        for line in lines:
            parts = line.split()
            if parts[1] == '0':
                f.write(parts[0] + '\n')


def write_result(file_list, ip_file, is_in, not_in,writeLine=-1):
    """
    Writes the result of IP address comparison to the specified Excel files.

    Args:
        file_list (list): A list of file paths to Excel files.
        ip_file (str): The path to the IP address file.
        is_in (str): The value to be written in the result column if the IP address is found in the IP address file.
        not_in (str): The value to be written in the result column if the IP address is not found in the IP address file.
    """
    for fn in file_list:
        try:
            wb = load_workbook(fn)
        except PermissionError:
            print(f"{fn}文件被占用,无法读取.")
            sys.exit(1)

        for sheet in wb.worksheets:
            col_num = get_ip_row_in_sheet(sheet)
            if col_num <= -1:
                continue
            max_col = sheet.max_column

            if writeLine == -1:
                writeLine = max_col + 1
            sheet.cell(row=1, column=writeLine).value = datetime.now().strftime("在线情况%m%d %H:%M")

            with open(ip_file, "r") as f:
                ip_list = set(f.read().splitlines())

            for row in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col_num).value  # sheet.cell 的 column 是从1开始

                if cell_value in ip_list:
                    sheet.cell(row=row, column=writeLine).value = is_in
                else:
                    sheet.cell(row=row, column=writeLine).value = not_in
        wb.save(fn)


def removeFile(fileList):
    for file in fileList:
        if Path(file).exists():
            Path(file).unlink()

def ip_xlsx_test(file_list, ping_timeout, ping_size, times, writeLine):
    """
    Perform IP testing using ping and write the results to files.

    Args:
        file_list (list): List of files to process.
        ping_timeout (int): Timeout value for ping in milliseconds.
        ping_size (int): Size of the ping packet in bytes.
        times (int): Number of times to repeat the ping operation.

    Returns:
        None
    """
    ip_list_name = "ip.txt"
    result_name = "result.txt"
    bad_name = "bad.txt"
    tempFileList = [Path.cwd().joinpath(ip_list_name), Path.cwd().joinpath(result_name), Path.cwd().joinpath(bad_name)]
    ping_relative_path = r'tools\PingInfoView.exe'
    get_ip_list(file_list, ip_list_name)
    # 未测试时bad ip为全部ip
    with open(ip_list_name, 'r') as file1, open(bad_name, 'w') as file2:
        file2.write(file1.read())
    # ping重复操作
    print(f'Ping Timeout={ping_timeout}ms {ping_size}Bytes.')
    
    try:
        while(times > 0):
            call_ping(ping_relative_path, bad_name, result_name, ping_timeout, ping_size)
            get_bad_ip(result_name, bad_name)
            times = times - 1
            print(f'{times}times left  ',end='')
    except KeyboardInterrupt:
        print("通过按键中断立即结束了ping操作")
    
    # 结果回写
    write_result(file_list, bad_name, "离线", "在线", writeLine)
    return tempFileList

def test_ip(relative_path=''):
    """
    使用ping检测IP地址并将结果写回相关文件.

    Args:
        relative_path (str): 包含需要检测的IP地址xlsx文件的相对路径.

    Returns:
        bool: True if the operation is successful, False otherwise.
    """
    locale.setlocale(locale.LC_ALL, 'en_US.utf8')
    config = configparser.ConfigParser()
    config.read('config/ping.ini', encoding="utf-8-sig")

    ping_timeout = config.getint('cfg', 'ping_timeout', fallback=2000)    #ping超时时间
    ping_times = config.getint('cfg', 'ping_times', fallback=10)        #ping尝试次数
    writeLine = config.getint('cfg', 'writeLine', fallback=-1)        #结果写回在xlsx的第几列
    ping_size = 32

    try:
        file_list = [relative_path + f for f in listdir(relative_path) if f.endswith('.xlsx') and '~' not in f]
        tempFileList = ip_xlsx_test(file_list, ping_timeout, ping_size, ping_times, writeLine)
    except Exception:
        traceback.print_exc()
        print("testip发生异常.")
        return False
    finally:
        removeFile(tempFileList)
    return True

if __name__ == '__main__':
    #test_ip(relative_path=".\数据文件\\")
    ip_list_name = "ip.txt"
    result_name = "result.txt"
    bad_name = "bad.txt"
    tempFileList = [ip_list_name, result_name, bad_name]
    removeFile(tempFileList)
    sys.exit(0)
